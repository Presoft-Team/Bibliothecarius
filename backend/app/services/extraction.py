from dataclasses import dataclass

import pytesseract
from pdf2image import convert_from_path
from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook

from app.core.config import settings

NATIVE = "native"
OCR_FALLBACK = "ocr_fallback"


@dataclass
class ExtractedPage:
    page_number: int
    text: str
    method: str
    confidence: float


def extract_document(file_path: str, content_type: str) -> list[ExtractedPage]:
    if content_type == "pdf":
        return _extract_pdf(file_path)
    if content_type == "docx":
        return _extract_docx(file_path)
    if content_type == "xlsx":
        return _extract_xlsx(file_path)
    if content_type == "txt":
        return _extract_txt(file_path)
    raise ValueError(f"Unsupported content type: {content_type}")


def _extract_pdf(file_path: str) -> list[ExtractedPage]:
    reader = PdfReader(file_path)
    pages: list[ExtractedPage] = []
    ocr_images = None  # lazily rendered only if a page actually needs OCR

    for i, page in enumerate(reader.pages):
        page_number = i + 1
        native_text = (page.extract_text() or "").strip()

        if len(native_text) >= settings.ocr_fallback_char_threshold:
            pages.append(ExtractedPage(page_number, native_text, NATIVE, confidence=1.0))
            continue

        if ocr_images is None:
            ocr_images = convert_from_path(file_path)
        ocr_text = pytesseract.image_to_string(ocr_images[i]).strip()
        # Still-empty OCR output means the page is likely blank rather than misdetected;
        # confidence stays low either way so the user can visually double-check it.
        confidence = 0.4 if ocr_text else 0.0
        pages.append(ExtractedPage(page_number, ocr_text, OCR_FALLBACK, confidence=confidence))

    return pages


def _extract_docx(file_path: str) -> list[ExtractedPage]:
    doc = DocxDocument(file_path)
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    # python-docx has no reliable page-boundary API, so the whole document is one "page".
    return [ExtractedPage(1, text, NATIVE, confidence=1.0)]


def _extract_xlsx(file_path: str) -> list[ExtractedPage]:
    workbook = load_workbook(file_path, data_only=True)
    pages: list[ExtractedPage] = []
    for i, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append("\t".join(cells))
        text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows)
        pages.append(ExtractedPage(i + 1, text, NATIVE, confidence=1.0))
    return pages


def _extract_txt(file_path: str) -> list[ExtractedPage]:
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        text = f.read()
    return [ExtractedPage(1, text, NATIVE, confidence=1.0)]
